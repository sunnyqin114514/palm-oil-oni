# -*- coding: utf-8 -*-
"""build_research_master_workbook.py — 生成导师要求的研究总表 Excel。

按导师五条修改意见：
  1. 数据整合到一张 Excel，交叉核验数据源
  2. 科学描绘季节性特征
  3. z-score 标准化
  4. 气象同步两两相关 + 表示函数
  5. 天气因子到单产领先 1-12 期相关 + 表示函数

输入：三本权威 Excel（不直接读 raw 文件）
输出：data/processed/archives/00_palm_oil_research_master_malaysia.xlsx
      code/model/figures/seasonality_*.png
      code/model/figures/lag_correlation_*.png
"""
from __future__ import annotations

import sys
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
plt.rcParams["font.sans-serif"] = ["PingFang SC", "Heiti SC", "Microsoft YaHei",
                                   "Arial Unicode MS", "Noto Sans CJK SC", "sans-serif"]
plt.rcParams["axes.unicode_minus"] = False
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from scipy import stats

MODEL_DIR = Path(__file__).resolve().parent
sys.path.insert(0, str(MODEL_DIR))
import data_pipeline  # noqa: E402

PROJECT_ROOT = MODEL_DIR.parent.parent
ARCHIVE_DIR = PROJECT_ROOT / "data" / "processed" / "archives"
FIGURES_DIR = MODEL_DIR / "figures"
FIGURES_DIR.mkdir(parents=True, exist_ok=True)

OUT_XLSX = ARCHIVE_DIR / "00_palm_oil_research_master_malaysia.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Microsoft YaHei", size=10)
README_TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E78")
README_KEY_FONT = Font(name="Microsoft YaHei", size=11, bold=True)


# ═══════════════════════════════════════════════════════════════════════════
# 通用工具
# ═══════════════════════════════════════════════════════════════════════════

def df_to_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    number_formats: Dict[str, str] | None = None,
) -> None:
    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        sample = df[col_name].astype(str).head(30).str.len().max() if len(df) else 6
        ws.column_dimensions[letter].width = min(max(int(sample or 6) + 2, 10), 36)
        if number_formats and col_name in number_formats:
            for row_idx in range(2, len(df) + 2):
                ws.cell(row=row_idx, column=col_idx).number_format = number_formats[col_name]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = BODY_FONT
    if len(df):
        ws.freeze_panes = "A2"


def add_readme(wb: Workbook, lines: List[str]) -> None:
    ws = wb.create_sheet("00_README", 0)
    ws.column_dimensions["A"].width = 90
    for line in lines:
        ws.append([line])
    ws.cell(row=1, column=1).font = README_TITLE_FONT
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).font = BODY_FONT


# ═══════════════════════════════════════════════════════════════════════════
# 1. 交叉核验
# ═══════════════════════════════════════════════════════════════════════════

def build_source_check(master: pd.DataFrame) -> pd.DataFrame:
    """对比主宽表与 data_pipeline 输出的关键统计量。"""
    pipeline_df = data_pipeline.build_dataset()

    checks: List[Dict[str, Any]] = []

    def add_check(metric: str, col: str, src_val: float, master_val: float):
        checks.append({
            "检验项": metric,
            "字段": col,
            "data_pipeline值": round(src_val, 6),
            "主宽表值": round(master_val, 6),
            "差值": round(master_val - src_val, 8),
            "通过": "✓" if abs(master_val - src_val) < 1e-4 else "✗",
        })

    add_check("行数", "全表", len(pipeline_df), len(master))
    for col in ["Yield", "Production", "Mature_Area", "PRCP", "TAVG", "ONI_lag12",
                "PRCP_DEV", "TAVG_DEV", "PRCP_DEV_3m", "TAVG_DEV_3m"]:
        if col in pipeline_df.columns and col in master.columns:
            add_check("均值", col, pipeline_df[col].mean(), master[col].mean())
            add_check("最小值", col, pipeline_df[col].min(), master[col].min())
            add_check("最大值", col, pipeline_df[col].max(), master[col].max())

    add_check("时间起点", "Date",
              float(pipeline_df["Date"].min()[:4]),
              float(master["Date"].min()[:4]))
    add_check("时间终点", "Date",
              float(pipeline_df["Date"].max()[:4]),
              float(master["Date"].max()[:4]))

    return pd.DataFrame(checks)


# ═══════════════════════════════════════════════════════════════════════════
# 2. 季节性分析
# ═══════════════════════════════════════════════════════════════════════════

SEASON_VARS = ["Yield", "Production", "PRCP", "TAVG", "ONI_lag12",
               "PRCP_DEV", "TAVG_DEV", "PRCP_DEV_3m", "TAVG_DEV_3m"]


def build_seasonality(master: pd.DataFrame) -> pd.DataFrame:
    rows: List[Dict[str, Any]] = []
    for var in SEASON_VARS:
        if var not in master.columns:
            continue
        for month in range(1, 13):
            subset = master.loc[master["Month"] == month, var].dropna()
            rows.append({
                "变量": var,
                "月份": month,
                "均值": round(subset.mean(), 6) if len(subset) else np.nan,
                "中位数": round(subset.median(), 6) if len(subset) else np.nan,
                "标准差": round(subset.std(), 6) if len(subset) else np.nan,
                "最小值": round(subset.min(), 6) if len(subset) else np.nan,
                "最大值": round(subset.max(), 6) if len(subset) else np.nan,
                "样本数": int(len(subset)),
            })
    return pd.DataFrame(rows)


def plot_seasonality(season_df: pd.DataFrame) -> List[Path]:
    """为每个变量画季节性折线图（均值 ± 1 标准差）。"""
    paths: List[Path] = []
    for var in season_df["变量"].unique():
        sub = season_df[season_df["变量"] == var].sort_values("月份")
        fig, ax = plt.subplots(figsize=(8, 4))
        ax.plot(sub["月份"], sub["均值"], "o-", color="#1f77b4", linewidth=2)
        ax.fill_between(
            sub["月份"],
            sub["均值"] - sub["标准差"],
            sub["均值"] + sub["标准差"],
            alpha=0.2, color="#1f77b4",
        )
        ax.set_xlabel("Month")
        ax.set_ylabel(var)
        ax.set_title(f"Seasonal Profile: {var}")
        ax.set_xticks(range(1, 13))
        ax.grid(True, alpha=0.3)
        fig.tight_layout()
        path = FIGURES_DIR / f"seasonality_{var}.png"
        fig.savefig(path, dpi=150)
        plt.close(fig)
        paths.append(path)
    return paths


# ═══════════════════════════════════════════════════════════════════════════
# 3. z-score 标准化（两种口径）
# ═══════════════════════════════════════════════════════════════════════════

ZSCORE_VARS = ["Yield", "ONI_lag12", "PRCP", "TAVG", "PRCP_DEV", "TAVG_DEV",
               "PRCP_DEV_3m", "TAVG_DEV_3m"]


def build_zscore(master: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """返回 (标准化宽表, 统计量表)。

    导师反馈第 3 条指出原版 z-score 只用全序列均值/方差做标准化，量纲虽被消除
    但年内季节性相位仍然存在，会让产量与气候因子的"领先 8-10 月相关"被季节周期
    放大成伪显著。这里同时给出两种标准化：

      * {var}_z   —— 全序列 z-score（去量纲，保留季节性）。用于跨变量量纲统一的横
                     截面比较，与原版一致。
      * {var}_mz  —— 按月份 z-score（在同一个月内做标准化）。这一步把"该月历史
                     平均水平"从序列里扣掉，等价于去季节距平的标准化形式，专门
                     用于领先期相关分析以排除季节相位重合。
    """
    result = master[["Date", "Month"]].copy()
    stats_rows: List[Dict[str, Any]] = []

    for var in ZSCORE_VARS:
        if var not in master.columns:
            continue
        col = master[var].dropna()
        mu = col.mean()
        sigma = col.std()
        z_col = f"{var}_z"
        mz_col = f"{var}_mz"
        result[z_col] = (master[var] - mu) / sigma if sigma > 0 else 0.0

        # 按月份做 z-score：先按 Month 分组求该月历史均值/标准差，再做标准化
        g = master.groupby("Month")[var]
        m_mean = g.transform("mean")
        m_std = g.transform("std")
        result[mz_col] = np.where(
            (m_std > 0) & master[var].notna(),
            (master[var] - m_mean) / m_std,
            0.0,
        )

        stats_rows.append({
            "变量": var,
            "全序列均值": round(mu, 6),
            "全序列标准差": round(sigma, 6),
            "全z均值": round(result[z_col].mean(), 8),
            "全z标准差": round(result[z_col].std(), 6),
            "按月z均值": round(result[mz_col].mean(), 8),
            "按月z标准差": round(result[mz_col].std(), 6),
            "样本数": int(col.count()),
        })

    return result, pd.DataFrame(stats_rows)


# ═══════════════════════════════════════════════════════════════════════════
# 4. 气象两两相关 + 表示函数
# ═══════════════════════════════════════════════════════════════════════════

WEATHER_PAIRS = [
    ("ONI_lag12", "PRCP_DEV_3m"),
    ("ONI_lag12", "TAVG_DEV_3m"),
    ("PRCP_DEV_3m", "TAVG_DEV_3m"),
    ("ONI_lag12", "PRCP"),
    ("ONI_lag12", "TAVG"),
    ("PRCP", "TAVG"),
]


def build_weather_correlation(master_with_z: pd.DataFrame) -> pd.DataFrame:
    """气象变量两两相关：原始值口径 + 按月份 z-score（去季节）口径并列。

    导师反馈第 2 条指出同步关系不显著。这里给出两个口径：
      * 原始值 Pearson r —— 含季节性，受相位重合干扰；
      * 按月份 z-score Pearson r —— 去掉年内季节性，反映气候因子之间真实的
        同步协动。若原始显著但去季节后塌缩，说明只是季节相位重合；若两个口径
        都不显著，说明在马来西亚范围内 ONI 与本地降水/温度的同步关系确实很弱，
        需要在模型里改用滞后项或区域聚合方案。
    """
    rows: List[Dict[str, Any]] = []
    for x_var, y_var in WEATHER_PAIRS:
        if x_var not in master_with_z.columns or y_var not in master_with_z.columns:
            continue
        xz = f"{x_var}_mz"
        yz = f"{y_var}_mz"

        valid_raw = master_with_z[[x_var, y_var]].dropna()
        n_raw = len(valid_raw)
        r_raw = p_raw = slope_raw = intercept_raw = np.nan
        if n_raw >= 10:
            r_raw, p_raw = stats.pearsonr(valid_raw[x_var], valid_raw[y_var])
            slope_raw, intercept_raw, _, _, _ = stats.linregress(valid_raw[x_var], valid_raw[y_var])

        r_mz = p_mz = slope_mz = intercept_mz = np.nan
        n_mz = 0
        if xz in master_with_z.columns and yz in master_with_z.columns:
            valid_mz = master_with_z[[xz, yz]].dropna()
            n_mz = len(valid_mz)
            if n_mz >= 10:
                r_mz, p_mz = stats.pearsonr(valid_mz[xz], valid_mz[yz])
                slope_mz, intercept_mz, _, _, _ = stats.linregress(valid_mz[xz], valid_mz[yz])

        def sig(p):
            if not np.isfinite(p):
                return ""
            return "***" if p < 0.001 else ("**" if p < 0.01 else ("*" if p < 0.05 else "ns"))

        rows.append({
            "X变量": x_var,
            "Y变量": y_var,
            "原始_Pearson_r": round(r_raw, 5) if np.isfinite(r_raw) else np.nan,
            "原始_p值": round(p_raw, 6) if np.isfinite(p_raw) else np.nan,
            "原始_显著性": sig(p_raw),
            "去季节_Pearson_r": round(r_mz, 5) if np.isfinite(r_mz) else np.nan,
            "去季节_p值": round(p_mz, 6) if np.isfinite(p_mz) else np.nan,
            "去季节_显著性": sig(p_mz),
            "去季节_斜率b": round(slope_mz, 6) if np.isfinite(slope_mz) else np.nan,
            "去季节_截距a": round(intercept_mz, 6) if np.isfinite(intercept_mz) else np.nan,
            "去季节表达式": (
                f"{y_var}_mz = {intercept_mz:.4f} + {slope_mz:.4f} × {x_var}_mz"
                if np.isfinite(slope_mz) else ""
            ),
            "原始样本数": n_raw,
            "去季节样本数": n_mz,
        })
    return pd.DataFrame(rows)


# ═══════════════════════════════════════════════════════════════════════════
# 5. 天气因子领先 1-12 期与单产相关（三口径对比，排除季节相位伪相关）
# ═══════════════════════════════════════════════════════════════════════════

LAG_FACTORS = ["ONI_lag12", "PRCP_DEV_3m", "TAVG_DEV_3m", "PRCP", "TAVG"]
MAX_LEAD = 12


def _lead_corr_one(master: pd.DataFrame, x_col: str, k: int, y_col: str = "Yield") -> Tuple[float, float, int]:
    """计算 x(t-k) 与 y(t) 的 Pearson r、p 值与样本数。"""
    shifted = master[["Date", x_col]].copy()
    shifted["Date"] = (
        pd.to_datetime(shifted["Date"]) + pd.DateOffset(months=k)
    ).dt.to_period("M").astype(str)
    shifted = shifted.rename(columns={x_col: "_x_lag"})
    merged = master[["Date", y_col]].merge(shifted, on="Date", how="inner").dropna()
    n = len(merged)
    if n < 20:
        return np.nan, np.nan, n
    r, p = stats.pearsonr(merged["_x_lag"], merged[y_col])
    return float(r), float(p), n


def build_yield_lag_correlation(master: pd.DataFrame) -> pd.DataFrame:
    """三口径并列：原始值、全序列 z-score、按月份 z-score（去季节）。

    导师反馈第 3 条要求排除"季节相位重合"导致的伪领先。对比三口径可以直接看出
    去掉季节性后哪些领先期仍然显著——只有按月份 z-score 列仍然 |r| 较大且 p<0.05
    的才是真信号；原始值列显著但按月 z 列塌缩到 0 的，属于季节相位重合伪相关。
    """
    rows: List[Dict[str, Any]] = []
    for var in LAG_FACTORS:
        if var not in master.columns:
            continue
        z_var = f"{var}_z"
        mz_var = f"{var}_mz"
        for k in range(1, MAX_LEAD + 1):
            r_raw, p_raw, n_raw = _lead_corr_one(master, var, k)
            r_z, p_z, _ = _lead_corr_one(master, z_var, k) if z_var in master.columns else (np.nan, np.nan, 0)
            r_mz, p_mz, _ = _lead_corr_one(master, mz_var, k) if mz_var in master.columns else (np.nan, np.nan, 0)

            def sig(p):
                if not np.isfinite(p):
                    return ""
                return "***" if p < 0.001 else ("**" if p < 0.01 else ("*" if p < 0.05 else "ns"))

            # 用按月份 z-score（去季节）口径求表示函数，这才是排除了季节相位后的真实弹性
            slope, intercept = np.nan, np.nan
            if mz_var in master.columns and np.isfinite(r_mz):
                shifted = master[["Date", mz_var]].copy()
                shifted["Date"] = (
                    pd.to_datetime(shifted["Date"]) + pd.DateOffset(months=k)
                ).dt.to_period("M").astype(str)
                shifted = shifted.rename(columns={mz_var: "_x_lag"})
                m = master[["Date", "Yield_mz"]].merge(shifted, on="Date", how="inner").dropna()
                if len(m) >= 20:
                    slope, intercept, _, _, _ = stats.linregress(m["_x_lag"], m["Yield_mz"])

            rows.append({
                "天气变量": var,
                "领先月数k": k,
                "原始_r": round(r_raw, 5) if np.isfinite(r_raw) else np.nan,
                "原始_p": round(p_raw, 6) if np.isfinite(p_raw) else np.nan,
                "原始_显著性": sig(p_raw),
                "全z_r": round(r_z, 5) if np.isfinite(r_z) else np.nan,
                "全z_p": round(p_z, 6) if np.isfinite(p_z) else np.nan,
                "全z_显著性": sig(p_z),
                "按月z_r(去季节)": round(r_mz, 5) if np.isfinite(r_mz) else np.nan,
                "按月z_p": round(p_mz, 6) if np.isfinite(p_mz) else np.nan,
                "按月z_显著性": sig(p_mz),
                "去季节表示函数_斜率b": round(slope, 6) if np.isfinite(slope) else np.nan,
                "去季节表示函数_截距a": round(intercept, 6) if np.isfinite(intercept) else np.nan,
                "样本数": n_raw,
            })
    return pd.DataFrame(rows)


def plot_lag_correlation(lag_df: pd.DataFrame) -> List[Path]:
    """为每个天气变量画三口径并列对比图。"""
    paths: List[Path] = []
    for var in lag_df["天气变量"].unique():
        sub = lag_df[lag_df["天气变量"] == var].sort_values("领先月数k")
        fig, ax = plt.subplots(figsize=(9, 4.5))
        x = sub["领先月数k"]
        width = 0.27
        ax.bar(x - width, sub["原始_r"], width, color="#9ecae1", label="原始值 r")
        ax.bar(x, sub["全z_r"], width, color="#fdae6b", label="全序列 z r")
        ax.bar(x + width, sub["按月z_r(去季节)"], width, color="#d62728", label="按月份 z r（去季节）")
        ax.axhline(0, color="black", linewidth=0.5)
        ax.set_xlabel("Lead months (k)")
        ax.set_ylabel("Pearson r with Yield")
        ax.set_title(f"Lead-Lag Correlation (3 口径对比): {var}(t-k) → Yield(t)")
        ax.set_xticks(range(1, MAX_LEAD + 1))
        ax.grid(True, alpha=0.3, axis="y")
        ax.legend(loc="best", fontsize=9)
        fig.tight_layout()
        path = FIGURES_DIR / f"lag_correlation_{var}.png"
        fig.savefig(path, dpi=150)
        plt.close(fig)
        paths.append(path)
    return paths


# ═══════════════════════════════════════════════════════════════════════════
# 主入口
# ═══════════════════════════════════════════════════════════════════════════

def main() -> None:
    print("[1/6] 构建主宽表...")
    master = data_pipeline.build_dataset()
    print(f"       rows={len(master)}, range={master['Date'].min()} ~ {master['Date'].max()}")
    n_backcast = (master["Area_Source"] == "BACKCAST").sum()
    n_mpob = (master["Area_Source"] == "MPOB").sum()
    print(f"       面积来源: MPOB={n_mpob} 行, BACKCAST={n_backcast} 行")

    print("[2/6] 交叉核验...")
    source_check = build_source_check(master)
    n_pass = (source_check["通过"] == "✓").sum()
    n_total = len(source_check)
    print(f"       {n_pass}/{n_total} 项通过")

    print("[3/6] 季节性分析...")
    season_df = build_seasonality(master)
    season_figs = plot_seasonality(season_df)
    print(f"       {len(season_figs)} 张图 → {FIGURES_DIR}")

    print("[4/6] z-score 标准化（双口径）...")
    zscore_df, zscore_stats = build_zscore(master)
    z_cols = [c for c in zscore_df.columns if c.endswith("_z") or c.endswith("_mz")]
    print(f"       标准化列: {z_cols}")

    # 把 zscore 列并回 master，供领先期分析使用按月份 z-score (去季节) 口径
    master_with_z = master.merge(zscore_df, on=["Date", "Month"], how="left")

    print("[5/6] 气象两两相关...")
    weather_corr = build_weather_correlation(master_with_z)
    print(f"       {len(weather_corr)} 组相关关系")

    print("[6/6] 领先期相关分析（三口径对比）...")
    lag_corr = build_yield_lag_correlation(master_with_z)
    lag_figs = plot_lag_correlation(lag_corr)
    print(f"       {len(lag_corr)} 组 · {len(lag_figs)} 张图")

    # ─── 写入 Excel ───
    print("\n写入 Excel...")
    wb = Workbook()
    wb.remove(wb.active)

    readme_lines = [
        "马来西亚棕榈油研究总表 — 导师二次修订版（样本扩展 + 去季节 z-score）",
        "",
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"脚本: code/model/build_research_master_workbook.py",
        f"数据口径: 仅马来西亚主模型（产量、成熟面积、ONI、MY降水、MY温度）",
        "",
        "本版相对上一版的关键改动:",
        "  1. 样本窗口前推：MPOB 成熟面积只有 2015-2025，对 2010-2014 用线性外推",
        "     补齐（Area_Source=BACKCAST 标记），样本量从 132 提升到约 190+，",
        "     为样本外验证留出更充裕窗口。",
        "  2. z-score 双口径：除原有全序列 z-score (去量纲) 外，新增按月份 z-score",
        "     (在每个月内标准化，等价于去季节距平标准化)，专门用于领先期分析。",
        "  3. 领前期相关三口径对比：原始值/全 z/按月 z，用于诊断 8-10 月领先期",
        "     显著是否来自季节相位重合。仅按月 z 列仍显著的领先期才视为真信号。",
        "  4. 气象两两相关升级为原始 vs 去季节双口径，解释同步关系不显著原因。",
        "",
        "Sheet 索引:",
        "  01_Source_Check   — 源文件交叉核验结果",
        "  02_Master_Monthly — 月度主宽表（含 Area_Source 标记）",
        "  03_Seasonality    — 按月份统计季节性（均值/中位数/标准差）",
        "  04_ZScore         — 双口径标准化序列（_z 全序列, _mz 按月份）",
        "  04_ZScore_Stats   — 双口径标准化统计量",
        "  05_Weather_Corr   — 气象变量两两相关（原始 vs 去季节）+ 表示函数",
        "  06_Lag_Corr       — 天气因子领先1-12期与单产相关（三口径对比）+ 去季节表示函数",
        "  07_Method         — 计算方法简述",
        "",
        "源文件映射:",
        "  01_palm_oil_planted_area_malaysia.xlsx → 成熟面积 (2015-2025 MPOB + 2010-2014 线性外推)",
        "  02_climate_data_malaysia.xlsx          → ONI, PRCP, TAVG, ECMWF",
        "  03_palm_oil_production_malaysia.xlsx   → CPO月度产量 (Production)",
        "",
        "关键公式:",
        "  Yield = Production / Mature_Area",
        "  PRCP_DEV = PRCP - 同月历史平均降水",
        "  TAVG_DEV = TAVG - 同月历史平均温度",
        "  PRCP_DEV_3m = 3个月滚动平均(PRCP_DEV)",
        "  TAVG_DEV_3m = 3个月滚动平均(TAVG_DEV)",
        "  全序列 z-score: z = (x - 全序列均值) / 全序列标准差  (只去量纲, 保留季节性)",
        "  按月份 z-score: z_m = (x - 该月历史均值) / 该月历史标准差  (去量纲 + 去季节性)",
        "",
        "面积外推说明:",
        "  MPOB 2015-2025 的 Mature_Area/Immature_Area/Total_Area 对 YEAR 做一元线性回归,",
        "  外推 2010-2014 年度值并展开为月频。所有外推值在 Area_Source 列标记为 BACKCAST,",
        "  便于后续敏感性分析时剔除或加权处理。",
    ]
    add_readme(wb, readme_lines)

    df_to_sheet(wb, "01_Source_Check", source_check)
    df_to_sheet(wb, "02_Master_Monthly", master)
    df_to_sheet(wb, "03_Seasonality", season_df)
    df_to_sheet(wb, "04_ZScore", zscore_df)
    df_to_sheet(wb, "04_ZScore_Stats", zscore_stats)
    df_to_sheet(wb, "05_Weather_Corr", weather_corr)
    df_to_sheet(wb, "06_Lag_Corr", lag_corr)

    method_lines = pd.DataFrame([
        {"步骤": "1", "内容": "从三本权威Excel读取原数据：产量、成熟面积、ONI、降水、温度"},
        {"步骤": "2", "内容": "MPOB成熟面积仅2015-2025，对2010-2014按YEAR线性外推补齐（Area_Source=BACKCAST）"},
        {"步骤": "3", "内容": "按月合并为宽表，计算 Yield = Production / Mature_Area"},
        {"步骤": "4", "内容": "计算同月气候常态(climatology)，求距平 DEV = 实际值 - 常态值"},
        {"步骤": "5", "内容": "计算3个月滚动距平 (PRCP_DEV_3m, TAVG_DEV_3m)"},
        {"步骤": "6", "内容": "z-score双口径: 全序列z只去量纲; 按月份z(去季节)在每个月内标准化, 同时去量纲与去季节性"},
        {"步骤": "7", "内容": "按月份(1-12)分组统计各变量均值/中位数/标准差 → 季节性剖面"},
        {"步骤": "8", "内容": "气象两两相关: 原始值 + 按月z双口径Pearson r + 一元线性回归 Y=a+bX"},
        {"步骤": "9", "内容": "领先期分析三口径对比: weather(t-k) 对 Yield(t), k=1..12, 分别用原始/全z/按月z计算r与p"},
        {"步骤": "10","内容": "判定规则: 仅按月z(去季节)口径仍 |r|较大且p<0.05的领先期视为真信号; 原始值显著但去季节后塌缩视为季节相位伪相关"},
        {"步骤": "11","内容": "表示函数对去季节口径做一元线性回归: Yield_mz = a + b×weather_mz(t-k), 记录r/p/样本数"},
    ])
    df_to_sheet(wb, "07_Method", method_lines)

    wb.save(OUT_XLSX)
    print(f"\n✓ 研究总表已保存: {OUT_XLSX}")
    print(f"✓ 图表目录: {FIGURES_DIR}")


if __name__ == "__main__":
    main()
