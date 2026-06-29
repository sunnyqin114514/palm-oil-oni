# -*- coding: utf-8 -*-
"""replace_master_monthly_with_raw.py — 把 research master 的 02_Master_Monthly
替换为"原始归档宽表"：按月份合并产量+面积+ONI+PRCP+TAVG 的原始值, 不加衍生列。

输入:
  data/processed/archives/01_palm_oil_planted_area_malaysia.xlsx  (MPOB 成熟面积, 年频)
  data/processed/archives/02_climate_data_malaysia.xlsx          (PRCP, TAVG, ONI)
  data/processed/archives/03_palm_oil_production_malaysia.xlsx   (iFinD 月度产量)
  data/processed/archives/00_palm_oil_research_master_malaysia.xlsx  (目标文件)

输出:
  覆盖写入 00_palm_oil_research_master_malaysia.xlsx 的 02_Master_Monthly Sheet,
  其他 Sheet 保持不变。

合并逻辑:
  - 以"产量月份"为主时间轴 (2007-01 ~ 2026-05)
  - 面积是年频, 按年份 ffill 到该年 12 个月
  - PRCP/TAVG/ONI 直接按月份 left join
  - 保留所有原始字段, 不计算 Yield/Trend/DEV 等衍生列
  - 时间轴覆盖产量全集, 气候/ONI 缺失的月份保留空值, 便于核查源数据边界
"""
from __future__ import annotations

from pathlib import Path
from typing import List

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
ARC = PROJECT_ROOT / "data" / "processed" / "archives"
TARGET = ARC / "00_palm_oil_research_master_malaysia.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Microsoft YaHei", size=10)


def build_raw_wide() -> pd.DataFrame:
    prod = pd.read_excel(ARC / "03_palm_oil_production_malaysia.xlsx",
                         sheet_name="Production_Monthly_iFinD")
    prod = prod.copy()
    prod["Date"] = pd.to_datetime(prod["DATE"]).dt.to_period("M").astype(str)
    prod = prod.rename(columns={
        "DATE": "Production_DATE_原始",
        "ID": "Production_ID",
        "INDEX_NAME": "Production_INDEX_NAME",
        "VALUE": "Production_吨",
        "RTIME": "Production_RTIME",
    })
    prod = prod.sort_values("Date").reset_index(drop=True)

    area = pd.read_excel(ARC / "01_palm_oil_planted_area_malaysia.xlsx",
                         sheet_name="01_MPOB_Mature_Area")
    area = area.copy().rename(columns={
        "YEAR": "Year",
        "MATURE_HECTARES": "Mature_Area_公顷",
        "IMMATURE_HECTARES": "Immature_Area_公顷",
        "TOTAL_HECTARES": "Total_Area_公顷",
    })
    area["Year"] = pd.to_numeric(area["Year"], errors="coerce").astype(int)
    # 年频展开为月频
    monthly_area: List[dict] = []
    for _, r in area.sort_values("Year").iterrows():
        y = int(r["Year"])
        for m in range(1, 13):
            monthly_area.append({
                "Date": f"{y:04d}-{m:02d}",
                "Year": y,
                "Mature_Area_公顷": r["Mature_Area_公顷"],
                "Immature_Area_公顷": r["Immature_Area_公顷"],
                "Total_Area_公顷": r["Total_Area_公顷"],
            })
    area_m = pd.DataFrame(monthly_area)

    prcp = pd.read_excel(ARC / "02_climate_data_malaysia.xlsx",
                         sheet_name="01_MY_PRCP_History")
    prcp = prcp.rename(columns={
        "YM": "Date",
        "PRCP": "PRCP_mm_per_month",
        "PRCP_GRID_COUNT": "PRCP_GRID_COUNT",
        "LAT_MIN": "PRCP_LAT_MIN", "LAT_MAX": "PRCP_LAT_MAX",
        "LON_MIN": "PRCP_LON_MIN", "LON_MAX": "PRCP_LON_MAX",
        "SOURCE": "PRCP_SOURCE", "PRCP_UNIT": "PRCP_UNIT",
    })

    tavg = pd.read_excel(ARC / "02_climate_data_malaysia.xlsx",
                         sheet_name="02_MY_TAVG_History")
    tavg = tavg.rename(columns={
        "YM": "Date",
        "TAVG_C": "TAVG_摄氏度",
        "TAVG_C_CLIMATOLOGY": "TAVG_CLIM_摄氏度",
        "TAVG_C_ANOMALY": "TAVG_ANOM_摄氏度",
        "TAVG_C_GRID_COUNT": "TAVG_GRID_COUNT",
        "LAT_MIN": "TAVG_LAT_MIN", "LAT_MAX": "TAVG_LAT_MAX",
        "LON_MIN": "TAVG_LON_MIN", "LON_MAX": "TAVG_LON_MAX",
        "SOURCE": "TAVG_SOURCE", "TAVG_UNIT": "TAVG_UNIT",
    })

    oni = pd.read_excel(ARC / "02_climate_data_malaysia.xlsx",
                        sheet_name="03_ONI_Monthly")
    oni = oni.rename(columns={"YM": "Date", "ONI": "ONI_值"})

    # 主时间轴：产量的所有月份
    base = prod[["Date"]].copy()
    out = (base
           .merge(area_m, on="Date", how="left")
           .merge(prcp, on="Date", how="left")
           .merge(tavg, on="Date", how="left")
           .merge(oni, on="Date", how="left")
           .merge(prod.drop(columns=["Date"]), on=base.index, how="left")
           .drop(columns=["key_0"]))
    # 把 Production 列挪到 Date 后面
    cols = list(out.columns)
    front = ["Date", "Production_吨", "Production_DATE_原始",
             "Mature_Area_公顷", "Immature_Area_公顷", "Total_Area_公顷",
             "Year", "ONI_值", "PRCP_mm_per_month", "TAVG_摄氏度"]
    rest = [c for c in cols if c not in front]
    out = out[front + rest]
    return out


def write_sheet(wb: Workbook_like, sheet_name: str, df: pd.DataFrame) -> None:
    if sheet_name in wb.sheetnames:
        idx = wb.sheetnames.index(sheet_name)
        del wb[sheet_name]
    ws = wb.create_sheet(sheet_name, idx)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        sample = df[col_name].astype(str).head(30).str.len().max() if len(df) else 6
        ws.column_dimensions[letter].width = min(max(int(sample or 6) + 2, 10), 32)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = BODY_FONT
    if len(df):
        ws.freeze_panes = "A2"


def main() -> None:
    print("[1/2] 构建原始归档宽表 ...")
    raw = build_raw_wide()
    print(f"  rows={len(raw)}, cols={len(raw.columns)}")
    print(f"  Date 范围: {raw['Date'].min()} ~ {raw['Date'].max()}")
    print(f"  列: {list(raw.columns)}")

    print("[2/2] 写回 00_palm_oil_research_master_malaysia.xlsx ...")
    wb = load_workbook(TARGET)
    write_sheet(wb, "02_Master_Monthly", raw)
    wb.save(TARGET)
    print(f"\n✓ 已替换 02_Master_Monthly, 文件: {TARGET}")
    print(f"  共 {len(raw)} 行 × {len(raw.columns)} 列 (原始归档宽表, 无衍生列)")


if __name__ == "__main__":
    main()
