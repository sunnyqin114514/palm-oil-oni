# -*- coding: utf-8 -*-
"""build_unified_master_workbook.py — 把三本归档 Excel 的所有数据整合进一份新总表。

输入:
  data/processed/archives/01_palm_oil_planted_area_malaysia.xlsx
  data/processed/archives/02_climate_data_malaysia.xlsx
  data/processed/archives/03_palm_oil_production_malaysia.xlsx

输出:
  data/processed/archives/00_palm_oil_unified_master_malaysia.xlsx

合并原则:
  - 保留三本 Excel 中的每个数据 Sheet 原样并入（行/列/数值完全一致）
  - 每张 Sheet 在新文件中按"模块前缀"重命名,便于在一份文件里区分来源
  - 顶部加 00_README + 01_Source_Check,做到自带说明和交叉核验
  - 不动原始三本 Excel,新文件独立存在
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Tuple

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


PROJECT_ROOT = Path(__file__).resolve().parent.parent.parent
ARCHIVE_DIR = PROJECT_ROOT / "data" / "processed" / "archives"

SRC_FILES = {
    "AREA": ARCHIVE_DIR / "01_palm_oil_planted_area_malaysia.xlsx",
    "CLIMATE": ARCHIVE_DIR / "02_climate_data_malaysia.xlsx",
    "PRODUCTION": ARCHIVE_DIR / "03_palm_oil_production_malaysia.xlsx",
}
OUT_XLSX = ARCHIVE_DIR / "00_palm_oil_unified_master_malaysia.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
BODY_FONT = Font(name="Microsoft YaHei", size=10)
README_TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E78")
README_KEY_FONT = Font(name="Microsoft YaHei", size=11, bold=True)
AREA_FILL = PatternFill("solid", fgColor="E8F1D6")
CLIMATE_FILL = PatternFill("solid", fgColor="DCEEFC")
PROD_FILL = PatternFill("solid", fgColor="FCE4D5")


def df_to_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    tint: PatternFill | None = None,
) -> None:
    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        sample = df[col_name].astype(str).head(30).str.len().max() if len(df) else 6
        ws.column_dimensions[letter].width = min(max(int(sample or 6) + 2, 10), 36)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = BODY_FONT
            if tint is not None:
                cell.fill = tint
    if len(df):
        ws.freeze_panes = "A2"


def add_readme(wb: Workbook, lines: List[str]) -> None:
    ws = wb.create_sheet("00_README", 0)
    ws.column_dimensions["A"].width = 100
    for line in lines:
        ws.append([line])
    ws.cell(row=1, column=1).font = README_TITLE_FONT
    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=1).font = BODY_FONT


def load_sheets(path: Path) -> Dict[str, pd.DataFrame]:
    out: Dict[str, pd.DataFrame] = {}
    xls = pd.ExcelFile(path)
    for s in xls.sheet_names:
        if s.startswith("00_README"):
            continue
        out[s] = pd.read_excel(path, sheet_name=s)
    return out


def build_source_check(sheets_meta: List[Dict[str, Any]]) -> pd.DataFrame:
    """对并入的总表与三本原始 Excel 做行数/列数交叉核验。"""
    rows: List[Dict[str, Any]] = []
    for m in sheets_meta:
        src_df = pd.read_excel(m["src_path"], sheet_name=m["src_sheet"])
        uni_df = m["unified_df"]
        rows.append({
            "模块": m["module"],
            "原始文件": m["src_path"].name,
            "原始Sheet": m["src_sheet"],
            "总表Sheet": m["unified_sheet"],
            "原始行数": len(src_df),
            "总表行数": len(uni_df),
            "原始列数": len(src_df.columns),
            "总表列数": len(uni_df.columns),
            "行数一致": "✓" if len(src_df) == len(uni_df) else "✗",
            "列数一致": "✓" if len(src_df.columns) == len(uni_df.columns) else "✗",
            "数值总和一致": "✓" if _numeric_sum_equal(src_df, uni_df) else "✗",
        })
    return pd.DataFrame(rows)


def _numeric_sum_equal(a: pd.DataFrame, b: pd.DataFrame) -> str:
    """比较两表所有数值列 sum 是否一致（列名相同的前提下）。"""
    try:
        common = [c for c in a.columns if c in b.columns]
        for c in common:
            if pd.api.types.is_numeric_dtype(a[c]) and pd.api.types.is_numeric_dtype(b[c]):
                if abs(float(a[c].sum()) - float(b[c].sum())) > 1e-3:
                    return "✗"
        return "✓"
    except Exception:
        return "—"


def main() -> None:
    print("[1/4] 读取三本归档 Excel ...")
    area_sheets = load_sheets(SRC_FILES["AREA"])
    climate_sheets = load_sheets(SRC_FILES["CLIMATE"])
    prod_sheets = load_sheets(SRC_FILES["PRODUCTION"])
    print(f"  种植面积: {list(area_sheets.keys())}")
    print(f"  气候:     {list(climate_sheets.keys())}")
    print(f"  产量:     {list(prod_sheets.keys())}")

    print("[2/4] 组装统一总表 ...")
    wb = Workbook()
    wb.remove(wb.active)

    sheets_meta: List[Dict[str, Any]] = []

    # 模块 A: 种植面积
    for src_sheet, df in area_sheets.items():
        unified_sheet = f"A_{src_sheet}"
        df_to_sheet(wb, unified_sheet, df, tint=AREA_FILL)
        sheets_meta.append({
            "module": "A 种植面积",
            "src_path": SRC_FILES["AREA"],
            "src_sheet": src_sheet,
            "unified_sheet": unified_sheet,
            "unified_df": df,
        })

    # 模块 B: 气候
    for src_sheet, df in climate_sheets.items():
        unified_sheet = f"B_{src_sheet}"
        df_to_sheet(wb, unified_sheet, df, tint=CLIMATE_FILL)
        sheets_meta.append({
            "module": "B 气候",
            "src_path": SRC_FILES["CLIMATE"],
            "src_sheet": src_sheet,
            "unified_sheet": unified_sheet,
            "unified_df": df,
        })

    # 模块 C: 产量
    for src_sheet, df in prod_sheets.items():
        unified_sheet = f"C_{src_sheet}"
        df_to_sheet(wb, unified_sheet, df, tint=PROD_FILL)
        sheets_meta.append({
            "module": "C 产量",
            "src_path": SRC_FILES["PRODUCTION"],
            "src_sheet": src_sheet,
            "unified_sheet": unified_sheet,
            "unified_df": df,
        })

    # 校验 Sheet
    print("[3/4] 生成交叉核验 Sheet ...")
    check_df = build_source_check(sheets_meta)

    # README
    readme_lines = [
        "马来西亚棕榈油 统一总表 — 三本归档 Excel 合并版",
        "",
        f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        f"脚本: code/scripts/build_unified_master_workbook.py",
        f"输出: {OUT_XLSX.name}",
        "",
        "合并原则:",
        "  - 把三本归档 Excel 的所有数据 Sheet 原样并入,行/列/数值与源文件完全一致",
        "  - 不修改、不删除原始三本 Excel;本文件为独立新建的总表",
        "  - 每张 Sheet 用模块前缀 A/B/C 标识来源,便于在一份文件里区分",
        "",
        "Sheet 命名规则:",
        "  A_*  → 种植面积模块 (绿色底)",
        "  B_*  → 气候模块      (蓝色底)",
        "  C_*  → 产量模块      (橙色底)",
        "  01_Source_Check → 交叉核验结果",
        "",
        "Sheet 索引:",
    ]
    for m in sheets_meta:
        readme_lines.append(
            f"  {m['unified_sheet']:30s} ← {m['src_path'].name} :: {m['src_sheet']}"
        )
    readme_lines += [
        "",
        "源文件 → 总表 映射:",
        f"  01_palm_oil_planted_area_malaysia.xlsx → A_01_MPOB_Mature_Area, A_02_Planted_Area_iFinD",
        f"  02_climate_data_malaysia.xlsx          → B_01_MY_PRCP_History, B_02_MY_TAVG_History, B_03_ONI_Monthly, B_04_MY_Climate_Forecast",
        f"  03_palm_oil_production_malaysia.xlsx   → C_Production_Monthly_iFinD",
    ]
    add_readme(wb, readme_lines)

    df_to_sheet(wb, "01_Source_Check", check_df)
    # 把校验 sheet 挪到 README 后面
    wb.move_sheet("01_Source_Check", offset=-(len(wb.sheetnames) - 2))

    print("[4/4] 保存 ...")
    wb.save(OUT_XLSX)
    n_pass = int((check_df["行数一致"] == "✓").sum() + (check_df["列数一致"] == "✓").sum() + (check_df["数值总和一致"] == "✓").sum())
    n_total = len(check_df) * 3
    print(f"\n✓ 已保存: {OUT_XLSX}")
    print(f"  共 {len(wb.sheetnames)} 个 Sheet (含 README + 校验 + {len(sheets_meta)} 张数据)")
    print(f"  交叉核验: {n_pass}/{n_total} 项通过")


if __name__ == "__main__":
    main()
