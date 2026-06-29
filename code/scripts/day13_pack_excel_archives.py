# -*- coding: utf-8 -*-
"""day13_pack_excel_archives.py — 把关键数据打包成三本 Excel 归档。

气候 Excel 严格按"每种数据只保留一份"的标准:
  - 马来西亚过去降水    → NASA POWER 区域月度主表
  - 马来西亚过去气温    → NASA POWER 区域月度主表
  - ONI                → NOAA 月度长表
  - 马来西亚气候预测    → ECMWF/Copernicus 网格长表

剔除项 (本次不收录):
  - ONI 季度宽表 / Legacy 手工 Excel
  - PRCP 站点区域平均 / PRCP+ONI 合并表
  - TAVG 月度聚合 / NASA POWER 区域平均月表
  - 产量 YoY% / 12M 滚动均值 / 周度估算

输出: data/processed/archives/
  01_palm_oil_planted_area_malaysia.xlsx
  02_climate_data_malaysia.xlsx
  03_palm_oil_production_malaysia.xlsx
"""
from __future__ import annotations

from datetime import datetime
from io import StringIO
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from numbers_parser import Document as NumbersDoc
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


HERE = Path(__file__).resolve().parent
PROJECT_ROOT = HERE.parent.parent
DATA_DIR = PROJECT_ROOT / "data"
OUT_DIR = DATA_DIR / "processed" / "archives"
OUT_DIR.mkdir(parents=True, exist_ok=True)

HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
HEADER_FONT = Font(name="Microsoft YaHei", size=11, bold=True, color="FFFFFF")
README_TITLE_FONT = Font(name="Microsoft YaHei", size=14, bold=True, color="1F4E78")
README_KEY_FONT = Font(name="Microsoft YaHei", size=11, bold=True)
BODY_FONT = Font(name="Microsoft YaHei", size=10)


# ────────────────────────────────────────────────────────────────────────────
# 通用写入工具
# ────────────────────────────────────────────────────────────────────────────

def df_to_sheet(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    column_widths: Optional[Dict[str, int]] = None,
    number_formats: Optional[Dict[str, str]] = None,
) -> None:
    ws = wb.create_sheet(sheet_name)
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
    for col_idx, col_name in enumerate(df.columns, start=1):
        letter = get_column_letter(col_idx)
        if column_widths and col_name in column_widths:
            ws.column_dimensions[letter].width = column_widths[col_name]
        else:
            sample = df[col_name].astype(str).head(30).str.len().max() if len(df) else 6
            ws.column_dimensions[letter].width = min(max(int(sample or 6) + 2, 10), 36)
        if number_formats and col_name in number_formats:
            for row_idx in range(2, len(df) + 2):
                ws.cell(row=row_idx, column=col_idx).number_format = number_formats[col_name]
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            cell.font = BODY_FONT
    if len(df):
        ws.freeze_panes = "A2"


def add_readme(
    wb: Workbook,
    title: str,
    summary_lines: List[str],
    sheets_table: List[Dict[str, Any]],
    source_files: List[Dict[str, Any]],
) -> None:
    ws = wb.create_sheet("00_README", 0)
    ws.column_dimensions["A"].width = 36
    ws.column_dimensions["B"].width = 90

    def write_kv(key: str, value: str) -> None:
        ws.append([key, value])
        ws.cell(row=ws.max_row, column=1).font = README_KEY_FONT
        ws.cell(row=ws.max_row, column=2).font = BODY_FONT

    ws.append([title])
    ws.cell(row=1, column=1).font = README_TITLE_FONT
    ws.append([])
    write_kv("生成时间", datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
    write_kv("脚本", "code/scripts/day13_pack_excel_archives.py")
    write_kv("收录原则", "每类数据只保留一份权威/主口径表, 剔除重复来源和重复口径")
    write_kv("覆盖范围", " / ".join(summary_lines))
    ws.append([])

    ws.append(["Sheet 索引"])
    ws.cell(row=ws.max_row, column=1).font = README_KEY_FONT
    ws.append(["Sheet 名称", "说明 / 行列数 / 时间范围"])
    for c in ws[ws.max_row]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    for item in sheets_table:
        ws.append([item["sheet"], item["desc"]])
    ws.append([])

    ws.append(["源文件 → 本 Excel 对应关系"])
    ws.cell(row=ws.max_row, column=1).font = README_KEY_FONT
    ws.append(["源文件", "对应 Sheet"])
    for c in ws[ws.max_row]:
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
    for item in source_files:
        ws.append([item["src"], item["mapping"]])

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=2):
        for cell in row:
            cell.alignment = Alignment(vertical="center", wrap_text=True)


# ────────────────────────────────────────────────────────────────────────────
# 源文件读取
# ────────────────────────────────────────────────────────────────────────────

def _read_noaa_oni_raw_ascii(path: Path) -> pd.DataFrame:
    """NOAA oni.ascii.txt 固定 4 列 (SEAS, YR, TOTAL, ANOM), 空白分隔。"""
    rows: List[Dict[str, Any]] = []
    with open(path, "r", encoding="utf-8") as f:
        next(f)
        for line in f:
            parts = line.split()
            if len(parts) != 4:
                continue
            rows.append({
                "SEAS": parts[0],
                "YR": int(parts[1]),
                "SST_TOTAL_C": float(parts[2]),
                "ONI_ANOM_C": float(parts[3]),
            })
    return pd.DataFrame(rows)


def _read_noaa_daily_temperature_csv(path: Path) -> Tuple[str, pd.DataFrame]:
    """NOAA station daily CSV: 第 1 行 "data", 第 2 行站点名, 第 3 行真正列名。"""
    with open(path, "r", encoding="utf-8") as f:
        lines = f.readlines()
    station = lines[1].strip().strip('"').rstrip(",")
    df = pd.read_csv(StringIO("".join(lines[2:])))
    df.columns = [c.strip() for c in df.columns]
    return station, df


def _read_numbers_table(path: Path, header_row_index: int = 1) -> Tuple[str, pd.DataFrame]:
    """读取 .numbers 文件第一张表, 返回 (站点描述, DataFrame)。"""
    doc = NumbersDoc(str(path))
    table = doc.sheets[0].tables[0]
    raw_rows = [[cell.value for cell in r] for r in table.rows()]
    header_meta = raw_rows[0][0] if (raw_rows and raw_rows[0]) else ""
    cols = [str(c) if c is not None else f"col_{i}" for i, c in enumerate(raw_rows[header_row_index])]
    df = pd.DataFrame(raw_rows[header_row_index + 1:], columns=cols)
    return str(header_meta or ""), df


def _read_nasa_power_monthly(path: Path, value_col: str) -> pd.DataFrame:
    """NASA POWER 网格 CSV 有元数据头, -END HEADER- 之后是 PARAMETER/YEAR/JAN..DEC 宽表。
    转成 (YM, LAT, LON, VALUE) 长表。这里只是"重排格式"而不做任何数值变换或聚合,
    保持每个 (网格点, 月份) 一行 = 原始观测粒度。"""
    with open(path, "r", encoding="utf-8") as f:
        body = f.read().split("-END HEADER-")[-1].strip()
    raw = pd.read_csv(StringIO(body))
    month_names = ["JAN", "FEB", "MAR", "APR", "MAY", "JUN", "JUL", "AUG", "SEP", "OCT", "NOV", "DEC"]
    month_map = {m: i + 1 for i, m in enumerate(month_names)}
    month_cols = [c for c in raw.columns if c.upper() in month_map]
    rows = []
    for _, row in raw.iterrows():
        year = int(row.get("YEAR"))
        lat = float(row.get("LAT"))
        lon = float(row.get("LON"))
        for m in month_cols:
            v = pd.to_numeric(row[m], errors="coerce")
            if pd.isna(v) or v == -999:
                continue
            rows.append({
                "YM": f"{year:04d}-{month_map[m.upper()]:02d}",
                "LAT": lat,
                "LON": lon,
                value_col: float(v),
            })
    return pd.DataFrame(rows).sort_values(["YM", "LAT", "LON"]).reset_index(drop=True)


# ────────────────────────────────────────────────────────────────────────────
# Excel 1 · 棕榈油种植面积
# ────────────────────────────────────────────────────────────────────────────

def build_planted_area_workbook() -> Tuple[Path, Dict[str, Any]]:
    # iFinD 总面积
    src_ifind = DATA_DIR / "raw" / "product" / "ifind_edb_palm_oil_planted_area_my.csv"
    raw_ifind = pd.read_csv(src_ifind)
    
    # MPOB 成熟/未成熟面积拆分
    src_mpob = DATA_DIR / "raw" / "product" / "mpob_palm_oil_planted_area_mature.csv"
    raw_mpob = pd.read_csv(src_mpob)

    wb = Workbook()
    wb.remove(wb.active)

    add_readme(
        wb,
        title="马来西亚 油棕种植面积 (Planted Area) — 原数据",
        summary_lines=[
            f"年度 {raw_mpob['YEAR'].min()}–{raw_mpob['YEAR'].max()}",
            f"{len(raw_mpob)} 行 · 单位: 公顷 (hectare)",
            "包含：总面积、成熟面积 (Mature)、未成熟面积 (Immature)",
        ],
        sheets_table=[
            {"sheet": "01_MPOB_Mature_Area", "desc": f"MPOB 官方报告提取的成熟/未成熟面积拆分, {len(raw_mpob)} 行"},
            {"sheet": "02_Planted_Area_iFinD", "desc": f"iFinD EDB 年度总面积原始返回, {len(raw_ifind)} 行"},
        ],
        source_files=[
            {"src": "data/raw/product/mpob_palm_oil_planted_area_mature.csv", "mapping": "→ 01_MPOB_Mature_Area (从 MPOB 官网 PDF 提取)"},
            {"src": "data/raw/product/ifind_edb_palm_oil_planted_area_my.csv", "mapping": "→ 02_Planted_Area_iFinD (iFinD S012222505 EDB 接口原始)"},
        ],
    )

    df_to_sheet(
        wb,
        "01_MPOB_Mature_Area",
        raw_mpob,
        number_formats={
            "MATURE_HECTARES": "#,##0",
            "IMMATURE_HECTARES": "#,##0",
            "TOTAL_HECTARES": "#,##0"
        },
    )

    df_to_sheet(
        wb,
        "02_Planted_Area_iFinD",
        raw_ifind,
        column_widths={"DATE": 14, "VALUE": 18, "ID": 14, "INDEX_NAME": 30, "RTIME": 22},
        number_formats={"VALUE": "#,##0"},
    )

    out = OUT_DIR / "01_palm_oil_planted_area_malaysia.xlsx"
    wb.save(out)
    return out, {"rows": len(raw_mpob)}


# ────────────────────────────────────────────────────────────────────────────
# Excel 2 · 气候 (ONI + PRCP + TAVG, 全部原数据)
# ────────────────────────────────────────────────────────────────────────────

def _read_oni_monthly_long() -> pd.DataFrame:
    df = pd.read_csv(DATA_DIR / "processed" / "meteo" / "noaa_nino34_pacific_oni_monthly.csv")
    df = df.rename(columns={c: "YM" for c in df.columns if c.lower() == "date"})
    df = df.rename(columns={c: "ONI" for c in df.columns if c.lower() in {"oni_value", "oni"}})
    return df[["YM", "ONI"]]


def _read_weather_monthly() -> pd.DataFrame:
    return pd.read_csv(DATA_DIR / "processed" / "meteo" / "nasa_power_malaysia_weather_monthly.csv")


def _read_ecmwf_seasonal_grid_csv() -> pd.DataFrame:
    return pd.read_csv(DATA_DIR / "processed" / "meteo" / "ecmwf_seasonal_forecast_malaysia_grid.csv")


def _read_historical_prcp() -> pd.DataFrame:
    df = _read_weather_monthly()
    return df[["YM", "PRCP", "PRCP_GRID_COUNT", "LAT_MIN", "LAT_MAX", "LON_MIN", "LON_MAX", "SOURCE", "PRCP_UNIT"]]


def _read_historical_tavg() -> pd.DataFrame:
    df = _read_weather_monthly()
    return df[["YM", "TAVG_C", "TAVG_C_CLIMATOLOGY", "TAVG_C_ANOMALY", "TAVG_C_GRID_COUNT", "LAT_MIN", "LAT_MAX", "LON_MIN", "LON_MAX", "SOURCE", "TAVG_UNIT"]]


def build_climate_workbook() -> Tuple[Path, Dict[str, Any]]:
    prcp_history = _read_historical_prcp()
    tavg_history = _read_historical_tavg()
    oni_monthly = _read_oni_monthly_long()
    ecmwf_forecast = _read_ecmwf_seasonal_grid_csv()

    wb = Workbook()
    wb.remove(wb.active)

    sheets_meta = [
        ("01_MY_PRCP_History", prcp_history,
         "马来西亚过去降水: NASA POWER 区域月度主表, mm/month"),
        ("02_MY_TAVG_History", tavg_history,
         "马来西亚过去气温: NASA POWER 区域月度主表, °C"),
        ("03_ONI_Monthly", oni_monthly,
         "ONI: NOAA Niño 3.4 月度长表"),
        ("04_MY_Climate_Forecast", ecmwf_forecast,
         "马来西亚气候预测: ECMWF/Copernicus 网格长表, 含集合成员与网格点"),
    ]

    add_readme(
        wb,
        title="马来西亚 气候数据 (每类仅保留一份)",
        summary_lines=[
            f"历史降水 {prcp_history['YM'].min()}–{prcp_history['YM'].max()}",
            f"历史气温 {tavg_history['YM'].min()}–{tavg_history['YM'].max()}",
            f"ONI {oni_monthly['YM'].min()}–{oni_monthly['YM'].max()}",
            f"ECMWF 季节预测 {ecmwf_forecast['FORECAST_REFERENCE_TIME'].min()} 起报 / 目标月 {ecmwf_forecast['TARGET_YM'].min()} / {len(ecmwf_forecast)} 行网格集合数据",
        ],
        sheets_table=[
            {"sheet": s, "desc": f"{d} | {len(df)} 行 × {len(df.columns)} 列"} for s, df, d in sheets_meta
        ],
        source_files=[
            {"src": "data/processed/meteo/nasa_power_malaysia_weather_monthly.csv", "mapping": "→ 01_MY_PRCP_History / 02_MY_TAVG_History"},
            {"src": "data/processed/meteo/noaa_nino34_pacific_oni_monthly.csv", "mapping": "→ 03_ONI_Monthly"},
            {"src": "data/processed/meteo/ecmwf_seasonal_forecast_malaysia_grid.csv", "mapping": "→ 04_MY_Climate_Forecast"},
        ],
    )

    for sheet, df, _desc in sheets_meta:
        df_to_sheet(wb, sheet, df)

    out = OUT_DIR / "02_climate_data_malaysia.xlsx"
    wb.save(out)
    return out, {
        "prcp_history_rows": len(prcp_history),
        "tavg_history_rows": len(tavg_history),
        "oni_monthly_rows": len(oni_monthly),
        "ecmwf_forecast_rows": len(ecmwf_forecast),
    }


# ────────────────────────────────────────────────────────────────────────────
# Excel 3 · 棕榈油产量
# ────────────────────────────────────────────────────────────────────────────

def build_production_workbook() -> Tuple[Path, Dict[str, Any]]:
    src = DATA_DIR / "raw" / "product" / "ifind_edb_palm_oil_production_my.csv"
    raw = pd.read_csv(src)

    wb = Workbook()
    wb.remove(wb.active)

    add_readme(
        wb,
        title="马来西亚 毛棕榈油 (CPO) 月度产量 — 原数据",
        summary_lines=[
            f"月度 {raw['DATE'].min()}–{raw['DATE'].max()}",
            f"{len(raw)} 行 · 单位: 吨",
            f"数据源: iFinD EDB S002958800 (来源: MPOB)",
        ],
        sheets_table=[
            {"sheet": "Production_Monthly_iFinD", "desc": f"iFinD EDB 月度原始返回, {len(raw)} 行, 字段未做任何数值变换"},
        ],
        source_files=[
            {"src": "data/raw/product/ifind_edb_palm_oil_production_my.csv", "mapping": "→ Production_Monthly_iFinD (iFinD S002958800 EDB 接口原始)"},
        ],
    )

    df_to_sheet(
        wb,
        "Production_Monthly_iFinD",
        raw,
        column_widths={"DATE": 14, "VALUE": 18, "ID": 14, "INDEX_NAME": 30, "RTIME": 22},
        number_formats={"VALUE": "#,##0"},
    )

    out = OUT_DIR / "03_palm_oil_production_malaysia.xlsx"
    wb.save(out)
    return out, {"rows": len(raw)}


# ────────────────────────────────────────────────────────────────────────────
# 入口
# ────────────────────────────────────────────────────────────────────────────

def main() -> None:
    p1, m1 = build_planted_area_workbook()
    p2, m2 = build_climate_workbook()
    p3, m3 = build_production_workbook()
    print(f"[Excel 1] {p1}\n          rows={m1['rows']}")
    print(f"[Excel 2] {p2}")
    for k, v in m2.items():
        print(f"          {k}={v}")
    print(f"[Excel 3] {p3}\n          rows={m3['rows']}")


if __name__ == "__main__":
    main()
